"""
Script để trích xuất dữ liệu từ các file HTM Strategy Tester Optimization
và tạo file Excel riêng cho từng file HTM.

Cách sử dụng:
    python htm_to_excel.py --input <folder_input> --output <folder_output>
    
Ví dụ:
    python htm_to_excel.py --input D:\\OptimizeLot --output D:\\Output
    
Mỗi file HTM sẽ được chuyển thành một file Excel riêng:
    audcad-144-sma.htm → audcad-144-sma.xlsx
    audusd-143.htm → audusd-143.xlsx
"""

import os
import argparse
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path


def extract_data_from_htm(file_path):
    """
    Trích xuất dữ liệu từ file HTM
    
    Args:
        file_path: Đường dẫn đến file HTM
        
    Returns:
        List các dictionary chứa dữ liệu từng dòng
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    soup = BeautifulSoup(content, 'html.parser')
    
    # Tìm tất cả các bảng
    tables = soup.find_all('table')
    
    # Bảng thứ 2 (index 1) chứa dữ liệu optimization
    if len(tables) < 2:
        print(f"Cảnh báo: File {file_path} không có đủ bảng dữ liệu")
        return []
    
    data_table = tables[1]
    
    # Lấy tất cả các hàng (bỏ qua hàng header)
    rows = data_table.find_all('tr')[1:]  # Bỏ hàng đầu tiên (header)
    
    results = []
    
    for row in rows:
        cells = row.find_all('td')
        if len(cells) < 7:
            continue
        
        try:
            # Lấy title từ attribute của cell Pass
            detail = cells[0].get('title', '')
            
            # Lấy giá trị Pass
            pass_value = cells[0].get_text(strip=True)
            
            # Lấy giá trị Profit
            profit = cells[1].get_text(strip=True)
            
            # Lấy Total trades
            total_trades = cells[2].get_text(strip=True)
            
            # Lấy Profit factor
            profit_factor = cells[3].get_text(strip=True)
            
            # Lấy Expected Payoff
            expected_payoff = cells[4].get_text(strip=True)
            
            # Lấy Drawdown $
            drawdown_dollar = cells[5].get_text(strip=True)
            
            # Lấy Drawdown %
            drawdown_percent = cells[6].get_text(strip=True)
            
            # Tính Profit/Drawdown$
            try:
                profit_val = float(profit.replace(',', ''))
                drawdown_val = float(drawdown_dollar.replace(',', ''))
                if drawdown_val > 0:
                    profit_drawdown = round(profit_val / drawdown_val, 2)
                else:
                    profit_drawdown = 0
            except:
                profit_drawdown = 0
            
            results.append({
                'Pass': pass_value,
                'Profit': profit,
                'Total trades': total_trades,
                'Profit factor': profit_factor,
                'Expected Payoff': expected_payoff,
                'Drawdown $': drawdown_dollar,
                'Drawdown %': drawdown_percent,
                'Profit/Drawdown$': profit_drawdown,
                'Detail': detail
            })
        except Exception as e:
            print(f"Lỗi khi xử lý dòng: {e}")
            continue
    
    return results


def create_excel_from_data(data, output_file, source_filename):
    """
    Tạo file Excel từ dữ liệu đã trích xuất
    
    Args:
        data: List các dictionary chứa dữ liệu
        output_file: Đường dẫn file Excel output
        source_filename: Tên file HTM gốc
    """
    # Tạo workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Optimization Results"
    
    # Tạo header
    headers = ['Pass', 'Profit', 'Total trades', 'Profit factor', 
               'Expected Payoff', 'Drawdown $', 'Drawdown %', 'Profit/Drawdown$', 'Detail']
    
    # Style cho header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Thêm dữ liệu
    current_row = 2
    for record in data:
        # Chuyển đổi sang số cho các cột number
        try:
            ws.cell(row=current_row, column=1).value = int(record['Pass'])
        except:
            ws.cell(row=current_row, column=1).value = record['Pass']
        
        try:
            ws.cell(row=current_row, column=2).value = float(record['Profit'].replace(',', ''))
        except:
            ws.cell(row=current_row, column=2).value = record['Profit']
        
        try:
            ws.cell(row=current_row, column=3).value = int(record['Total trades'])
        except:
            ws.cell(row=current_row, column=3).value = record['Total trades']
        
        try:
            ws.cell(row=current_row, column=4).value = float(record['Profit factor'])
        except:
            ws.cell(row=current_row, column=4).value = record['Profit factor']
        
        try:
            ws.cell(row=current_row, column=5).value = float(record['Expected Payoff'])
        except:
            ws.cell(row=current_row, column=5).value = record['Expected Payoff']
        
        try:
            ws.cell(row=current_row, column=6).value = float(record['Drawdown $'].replace(',', ''))
        except:
            ws.cell(row=current_row, column=6).value = record['Drawdown $']
        
        try:
            ws.cell(row=current_row, column=7).value = float(record['Drawdown %'])
        except:
            ws.cell(row=current_row, column=7).value = record['Drawdown %']
        
        # Profit/Drawdown$ đã là số rồi
        ws.cell(row=current_row, column=8).value = record['Profit/Drawdown$']
        
        # Detail là text
        ws.cell(row=current_row, column=9).value = record['Detail']
        
        current_row += 1
    
    # Tự động điều chỉnh độ rộng cột
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        # Giới hạn độ rộng cột Detail
        if column == 'I':  # Cột Detail
            adjusted_width = min(max_length + 2, 80)
        else:
            adjusted_width = min(max_length + 2, 30)
        
        ws.column_dimensions[column].width = adjusted_width
    
    # Lưu file
    wb.save(output_file)


def process_folder(input_folder, output_folder):
    """
    Xử lý tất cả file HTM trong folder và tạo file Excel tương ứng
    
    Args:
        input_folder: Đường dẫn folder chứa file HTM
        output_folder: Đường dẫn folder output để lưu các file Excel
    """
    input_path = Path(input_folder)
    
    if not input_path.exists():
        print(f"Lỗi: Folder {input_folder} không tồn tại!")
        return
    
    # Tìm tất cả file .htm và .html
    htm_files = list(input_path.glob('*.htm')) + list(input_path.glob('*.html'))
    
    if not htm_files:
        print(f"Không tìm thấy file HTM nào trong folder {input_folder}")
        return
    
    print(f"Tìm thấy {len(htm_files)} file HTM")
    
    # Tạo folder output nếu chưa có
    output_path = Path(output_folder)
    output_path.mkdir(parents=True, exist_ok=True)
    
    # Xử lý từng file
    total_files = 0
    total_records = 0
    
    for htm_file in htm_files:
        print(f"Đang xử lý: {htm_file.name}")
        
        # Trích xuất dữ liệu
        data = extract_data_from_htm(htm_file)
        
        if not data:
            print(f"  → Bỏ qua: Không có dữ liệu")
            continue
        
        # Tạo tên file Excel (thay đổi đuôi từ .htm thành .xlsx)
        excel_filename = htm_file.stem + '.xlsx'
        excel_path = output_path / excel_filename
        
        # Tạo file Excel
        create_excel_from_data(data, excel_path, htm_file.name)
        
        print(f"  → Đã tạo: {excel_filename} ({len(data)} dòng)")
        total_files += 1
        total_records += len(data)
    
    print(f"\n{'='*60}")
    print(f"Hoàn thành!")
    print(f"Tổng số file Excel đã tạo: {total_files}")
    print(f"Tổng số dòng dữ liệu: {total_records}")
    print(f"Folder output: {output_folder}")
    print(f"{'='*60}")


def main():
    """
    Hàm main để xử lý command line arguments
    """
    parser = argparse.ArgumentParser(
        description='Trích xuất dữ liệu từ các file HTM Strategy Tester và tạo file Excel'
    )
    
    parser.add_argument(
        '--input',
        '-i',
        required=True,
        help='Đường dẫn folder chứa các file HTM'
    )
    
    parser.add_argument(
        '--output',
        '-o',
        required=True,
        help='Đường dẫn folder output để lưu các file Excel'
    )
    
    args = parser.parse_args()
    
    process_folder(args.input, args.output)


if __name__ == '__main__':
    main()
